"""Implementation of analysis of ethnicities.
Uses a supplemental data excel spreadsheet to
calculate percentages of different hispanic
ethnicities from a list of hispanic registered
voters from a specified county.
"""

from openpyxl import Workbook, load_workbook
from argparse import ArgumentParser
from math import fsum



# set up argument parser
parser = ArgumentParser()
parser.add_argument('--countyData', required=True)
parser.add_argument('--countyName', required=True)
parser.add_argument('--supplementalData', required=True)
args = parser.parse_args()

# initialize workbook with openpyxl
wb = Workbook()
wb = load_workbook(filename=args.countyData, data_only=True)

# enter sheet of input county
county_sheet = wb[args.countyName]

# initialize workbook with supplemental data
dataWb = Workbook()
dataWb = load_workbook(filename=args.supplementalData, data_only=True)
# access to individual worksheets
firstNamesWs = dataWb['FirstN']
lastNamesWs = dataWb['LastN']
zipCodeWs = dataWb['ZIP']

#output workbook
outputWb = Workbook()
outputSheet = outputWb.active
outputSheet.title = 'Output'

# row to start at for saving to output worksheet
activeRow = 2

# dictionary for each sub ethnicity label
ethnicDict = {
    0 : 'Mexican',
    1 : 'PuertoRican',
    2 : 'Cuban',
    3 : 'Dominican',
    4 : 'CostaRican',
    5 : 'Guatemalan',
    6 : 'Honduran',
    7 : 'Nicaraguan',
    8 : 'Panamanian',
    9 : 'Salvadoran',
    10 : 'Argentinian',
    11 : 'Bolivian',
    12 : 'Chilean',
    13 : 'Colombian',
    14 : 'Ecuadorian',
    15 : 'Paraguayan',
    16 : 'Peruvian',
    17 : 'Uruguayan',
    18 : 'Venezuelan',
}

# First Row Labels for each output worksheet column
outputSheet.cell(row=1, column=1).value = 'County'
outputSheet.cell(row=1, column=2).value = 'Voter ID'
outputSheet.cell(row=1, column=3).value = 'Last Name'
outputSheet.cell(row=1, column=4).value = 'First Name'
outputSheet.cell(row=1, column=5).value = 'Middle Name'
outputSheet.cell(row=1, column=6).value = 'Zip Code'
outputSheet.cell(row=1, column=7).value = 'Gender'
outputSheet.cell(row=1, column=8).value = 'Race'
outputSheet.cell(row=1, column=9).value = 'Birth Date'
outputSheet.cell(row=1, column=10).value = 'Party'
outputSheet.cell(row=1, column=11).value = 'Voter Status'
outputSheet.cell(row=1, column=12).value = 'Sub Ethnicity'
outputSheet.cell(row=1, column=32).value = 'Others'
for i in range(0,19):
    outputSheet.cell(row=1, column=i+13).value = ethnicDict[i]

#create tuple data structure
initialPercentages = ([], [], [], [])
#format: (last, first, zip, secondLast)

# iterate through rows of people
for person in county_sheet.iter_rows(min_row=2): #min_row=2 skips label row
    
    first = ''
    last = ''
    secondLast = ''
    
    # parse last name and check if two last names
    # if for some reason someone has a multi part last name
    # with spaces that is considered a single last name
    # this will not work and consider it as two last names
    if person[2].value is not None:
        lastParse = person[2].value.replace('-', ' ').split()
        last = lastParse[0]
        if len(lastParse) > 1:
            secondLast = lastParse[1]
    # parse first name to take only first part
    if person[3].value is not None:
        firstParse = person[3].value.replace('-', ' ').split()
        first = firstParse[0]
    # get zip code first 5 digits
    zipCode = str(person[5].value)
    zipCode = zipCode[:5]
    
    # setting up variables
    percentage = 0.0
    others = 0.0
    
#    # set up tuple data structure
#    if len(lastParse) > 1:
#        initialPercentages = ([], [], [], [])
#        #format: (last, first, zip, secondLast)
#    else:
#        initialPercentages = ([], [], [])
#        #format: (last, first, zip)
    
    #clear lists in tuple
    initialPercentages[0].clear()
    initialPercentages[1].clear()
    initialPercentages[2].clear()
    initialPercentages[3].clear()
    
    # getting percentage from supplemental data
    
    # last name
    for i in range(0, 19):
        for row in lastNamesWs.iter_rows(min_row=2, min_col=1 + (2 * i), max_col=2 + (2 * i)):
            if row[0].value != 'None' and row[0].value == last:
                percentage = row[1].value
                break
            else:
                percentage = 0.0
        initialPercentages[0].append(percentage)
    
    # first name
    for i in range(0, 19):
        for row in firstNamesWs.iter_rows(min_row=2, min_col=1 + (2 * i), max_col=2 + (2 * i)):
            if row[0].value != 'None' and row[0].value == first:
                percentage = row[1].value
                break
            else:
                percentage = 0.0
        initialPercentages[1].append(percentage)
    
    # zip codes
    for i in range(0, 19):
        for row in zipCodeWs.iter_rows():
            if row[0].value != 'None' and str(row[0].value) == zipCode:
                percentage = row[6 + (i * 2)].value
                # grab value for other ethnicities
                others = row[44].value
                break
            else:
                percentage = 0.0
        initialPercentages[2].append(percentage)
    
    # second last (if needed) 
    if len(lastParse) > 1:
        for i in range(0, 19):
            for row in lastNamesWs.iter_rows(min_row=2, min_col=1 + (2 * i), max_col=2 + (2 * i)):
                if row[0].value != 'None' and row[0].value == secondLast:
                    percentage = row[1].value
                    break
                else:
                    percentage = 0.0
            initialPercentages[3].append(percentage)
    
    #if sum > 0 then use value else 1
    
    lastName = 0.0
    firstName = 0.0
    zips = 0.0
    secondLastName = 0.0
    
    # calculate percentage of sub ethnicities
    subEthnicities = []
    for i in range(0, 19):
        # first name
        if fsum(initialPercentages[0]) > 0:
            lastName = initialPercentages[0][i] * .7
        else:
            lastName = 1.0
        
        # last name
        if fsum(initialPercentages[1]) > 0:
            firstName = initialPercentages[1][i] * .3
        else:
            firstName = 1.0
        
        # zip code
        if fsum(initialPercentages[2]) > 0:
            zips = initialPercentages[2][i]
        
        # second last name
        if len(lastParse) > 1:
            if fsum(initialPercentages[3]) > 0:
                secondLastName = initialPercentages[3][i]
            else:
                secondLastName = 1.0
            # calculate percentage with second last name
            subEthnicities.append(lastName * firstName * zips * secondLastName)
        else:
            # calculate percentage with one last name
            subEthnicities.append(lastName * firstName * zips)
    
    # protect against divide by 0
    if fsum(subEthnicities) != 0 and person[5].value is not None:
        # Zip, Party, Gender, and Calculated Ethnicity of each person
        outputSheet.cell(row=activeRow, column=1).value = person[0].value
        outputSheet.cell(row=activeRow, column=2).value = person[1].value
        outputSheet.cell(row=activeRow, column=3).value = person[2].value
        outputSheet.cell(row=activeRow, column=4).value = person[3].value
        outputSheet.cell(row=activeRow, column=5).value = person[4].value
        outputSheet.cell(row=activeRow, column=6).value = int(str(person[5].value)[:5])
        outputSheet.cell(row=activeRow, column=7).value = person[6].value
        outputSheet.cell(row=activeRow, column=8).value = person[7].value
        outputSheet.cell(row=activeRow, column=9).value = person[8].value
        outputSheet.cell(row=activeRow, column=9).number_format = 'mm/dd/yyyy'
        outputSheet.cell(row=activeRow, column=10).value = person[9].value
        outputSheet.cell(row=activeRow, column=11).value = person[10].value
        maxPos = subEthnicities.index(max(subEthnicities))
        outputSheet.cell(row=activeRow, column=12).value = ethnicDict[maxPos]
        
        percentageSum = 0.0
        
        # Add percentages to each ethnicity column
        for i in range(0, 19):
            # final step
            calculatedPercentage = subEthnicities[i] / (fsum(subEthnicities) / (1 - others))
            outputSheet.cell(row=activeRow, column=i + 13).value = calculatedPercentage
            outputSheet.cell(row=activeRow, column=i + 13).number_format = '0%'
            percentageSum += calculatedPercentage
        
        # others column
        outputSheet.cell(row=activeRow, column=32).value = 1 - percentageSum
        outputSheet.cell(row=activeRow, column=32).number_format = '0%'
        
        # increment to next row
        activeRow += 1

# save workbook
outputFileName = args.countyName + '_output.xlsx'
outputWb.save(outputFileName)

#wb.close()
#dataWb.close()